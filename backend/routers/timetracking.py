"""
Time Tracking API Router
Allows admin to track work hours - both manual entries and auto-tracked sessions
"""

from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel, Field
from typing import Optional, List
from datetime import datetime, timedelta, timezone
from bson import ObjectId
import os
import io

router = APIRouter(prefix="/api/timetracking", tags=["Time Tracking"])

# MongoDB connection
from pymongo import MongoClient
MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "radiocheck")
client = MongoClient(MONGO_URL)
db = client[DB_NAME]
time_entries = db.time_entries
active_sessions = db.time_tracking_sessions

# Categories for time tracking with default hourly rates
CATEGORIES = [
    "Development",
    "Admin Portal",
    "Staff Portal", 
    "LMS Admin",
    "LMS Learning",
    "App Testing",
    "Support",
    "Training",
    "Documentation",
    "Meetings",
    "Other"
]

# Default hourly rate in GBP
DEFAULT_HOURLY_RATE = 35.00

# Category-specific rates (optional - defaults to DEFAULT_HOURLY_RATE)
CATEGORY_RATES = {
    "Development": 35.00,
    "Admin Portal": 35.00,
    "Staff Portal": 35.00,
    "LMS Admin": 35.00,
    "LMS Learning": 35.00,
    "App Testing": 35.00,
    "Support": 30.00,
    "Training": 30.00,
    "Documentation": 30.00,
    "Meetings": 25.00,
    "Other": 30.00
}

def get_rate_for_category(category: str) -> float:
    """Get the hourly rate for a category"""
    return CATEGORY_RATES.get(category, DEFAULT_HOURLY_RATE)

def calculate_cost(hours: int, minutes: int, category: str) -> float:
    """Calculate cost for a time entry"""
    rate = get_rate_for_category(category)
    total_hours = hours + (minutes / 60)
    return round(total_hours * rate, 2)

# Pydantic Models
class TimeEntryCreate(BaseModel):
    date: str = Field(..., description="Date in YYYY-MM-DD format")
    hours: int = Field(..., ge=0, le=24)
    minutes: int = Field(0, ge=0, le=59)
    category: str = Field(..., description="Work category")
    description: str = Field(..., description="What was done")
    
class TimeEntryUpdate(BaseModel):
    date: Optional[str] = None
    hours: Optional[int] = Field(None, ge=0, le=24)
    minutes: Optional[int] = Field(None, ge=0, le=59)
    category: Optional[str] = None
    description: Optional[str] = None

class SessionStart(BaseModel):
    category: str
    page_url: Optional[str] = None

class SessionEnd(BaseModel):
    session_id: str
    description: Optional[str] = None

def serialize_entry(entry):
    """Convert MongoDB document to JSON-serializable dict"""
    hours = entry.get("hours", 0)
    minutes = entry.get("minutes", 0)
    category = entry.get("category", "Other")
    cost = calculate_cost(hours, minutes, category)
    rate = get_rate_for_category(category)
    
    return {
        "id": str(entry["_id"]),
        "date": entry.get("date"),
        "hours": hours,
        "minutes": minutes,
        "category": category,
        "description": entry.get("description"),
        "auto_tracked": entry.get("auto_tracked", False),
        "hourly_rate": rate,
        "cost": cost,
        "created_at": entry.get("created_at").isoformat() if entry.get("created_at") else None,
        "updated_at": entry.get("updated_at").isoformat() if entry.get("updated_at") else None,
    }

# ============ CRUD Endpoints ============

@router.get("/categories")
async def get_categories():
    """Get available time tracking categories with their rates"""
    return {
        "categories": CATEGORIES,
        "rates": CATEGORY_RATES,
        "default_rate": DEFAULT_HOURLY_RATE
    }

@router.get("/rates")
async def get_rates():
    """Get hourly rates for all categories"""
    return {
        "rates": CATEGORY_RATES,
        "default_rate": DEFAULT_HOURLY_RATE
    }

@router.post("/entries")
async def create_entry(entry: TimeEntryCreate):
    """Create a manual time entry"""
    if entry.category not in CATEGORIES:
        raise HTTPException(status_code=400, detail=f"Invalid category. Must be one of: {CATEGORIES}")
    
    now = datetime.now(timezone.utc)
    doc = {
        "date": entry.date,
        "hours": entry.hours,
        "minutes": entry.minutes,
        "category": entry.category,
        "description": entry.description,
        "auto_tracked": False,
        "created_at": now,
        "updated_at": now,
    }
    
    result = time_entries.insert_one(doc)
    doc["_id"] = result.inserted_id
    
    return {"message": "Entry created", "entry": serialize_entry(doc)}

@router.get("/entries")
async def get_entries(
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    category: Optional[str] = Query(None, description="Filter by category"),
    limit: int = Query(100, le=500),
    skip: int = Query(0, ge=0)
):
    """Get time entries with optional filters"""
    query = {}
    
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    if category:
        query["category"] = category
    
    cursor = time_entries.find(query).sort("date", -1).skip(skip).limit(limit)
    entries = [serialize_entry(e) for e in cursor]
    
    total = time_entries.count_documents(query)
    
    return {
        "entries": entries,
        "total": total,
        "limit": limit,
        "skip": skip
    }

@router.get("/entries/{entry_id}")
async def get_entry(entry_id: str):
    """Get a specific time entry"""
    try:
        entry = time_entries.find_one({"_id": ObjectId(entry_id)})
    except:
        raise HTTPException(status_code=400, detail="Invalid entry ID")
    
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    return {"entry": serialize_entry(entry)}

@router.put("/entries/{entry_id}")
async def update_entry(entry_id: str, update: TimeEntryUpdate):
    """Update a time entry"""
    try:
        entry = time_entries.find_one({"_id": ObjectId(entry_id)})
    except:
        raise HTTPException(status_code=400, detail="Invalid entry ID")
    
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    
    if "category" in update_data and update_data["category"] not in CATEGORIES:
        raise HTTPException(status_code=400, detail=f"Invalid category")
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc)
        time_entries.update_one({"_id": ObjectId(entry_id)}, {"$set": update_data})
    
    updated = time_entries.find_one({"_id": ObjectId(entry_id)})
    return {"message": "Entry updated", "entry": serialize_entry(updated)}

@router.delete("/entries/{entry_id}")
async def delete_entry(entry_id: str):
    """Delete a time entry"""
    try:
        result = time_entries.delete_one({"_id": ObjectId(entry_id)})
    except:
        raise HTTPException(status_code=400, detail="Invalid entry ID")
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    return {"message": "Entry deleted"}

# ============ Session Tracking ============

@router.post("/session/start")
async def start_session(session: SessionStart):
    """Start an auto-tracking session"""
    now = datetime.now(timezone.utc)
    
    doc = {
        "category": session.category,
        "page_url": session.page_url,
        "started_at": now,
        "last_activity": now,
        "active": True,
    }
    
    result = active_sessions.insert_one(doc)
    
    return {
        "session_id": str(result.inserted_id),
        "started_at": now.isoformat()
    }

@router.post("/session/heartbeat/{session_id}")
async def session_heartbeat(session_id: str):
    """Update session activity timestamp"""
    try:
        result = active_sessions.update_one(
            {"_id": ObjectId(session_id), "active": True},
            {"$set": {"last_activity": datetime.now(timezone.utc)}}
        )
    except:
        raise HTTPException(status_code=400, detail="Invalid session ID")
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Active session not found")
    
    return {"message": "Heartbeat recorded"}

@router.post("/session/end")
async def end_session(data: SessionEnd):
    """End a session and create time entry"""
    try:
        session = active_sessions.find_one({"_id": ObjectId(data.session_id)})
    except:
        raise HTTPException(status_code=400, detail="Invalid session ID")
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Calculate duration
    started_at = session["started_at"]
    ended_at = datetime.now(timezone.utc)
    duration = ended_at - started_at
    
    total_minutes = int(duration.total_seconds() / 60)
    hours = total_minutes // 60
    minutes = total_minutes % 60
    
    # Only log if more than 1 minute
    if total_minutes >= 1:
        entry_doc = {
            "date": started_at.strftime("%Y-%m-%d"),
            "hours": hours,
            "minutes": minutes,
            "category": session["category"],
            "description": data.description or f"Auto-tracked session on {session.get('page_url', session['category'])}",
            "auto_tracked": True,
            "session_id": data.session_id,
            "created_at": ended_at,
            "updated_at": ended_at,
        }
        time_entries.insert_one(entry_doc)
    
    # Mark session as inactive
    active_sessions.update_one(
        {"_id": ObjectId(data.session_id)},
        {"$set": {"active": False, "ended_at": ended_at}}
    )
    
    return {
        "message": "Session ended",
        "duration_minutes": total_minutes,
        "hours": hours,
        "minutes": minutes
    }

# ============ Summary & Reports ============

@router.get("/summary")
async def get_summary(
    month: Optional[str] = Query(None, description="Month in YYYY-MM format"),
    year: Optional[int] = Query(None, description="Year for summary")
):
    """Get time tracking summary"""
    now = datetime.now(timezone.utc)
    
    # Determine date range
    if month:
        year_num, month_num = map(int, month.split("-"))
        start_date = f"{year_num}-{month_num:02d}-01"
        if month_num == 12:
            end_date = f"{year_num + 1}-01-01"
        else:
            end_date = f"{year_num}-{month_num + 1:02d}-01"
    elif year:
        start_date = f"{year}-01-01"
        end_date = f"{year + 1}-01-01"
    else:
        # Default to current month
        start_date = f"{now.year}-{now.month:02d}-01"
        if now.month == 12:
            end_date = f"{now.year + 1}-01-01"
        else:
            end_date = f"{now.year}-{now.month + 1:02d}-01"
    
    # Aggregate by category
    pipeline = [
        {"$match": {"date": {"$gte": start_date, "$lt": end_date}}},
        {"$group": {
            "_id": "$category",
            "total_hours": {"$sum": "$hours"},
            "total_minutes": {"$sum": "$minutes"},
            "entry_count": {"$sum": 1}
        }}
    ]
    
    results = list(time_entries.aggregate(pipeline))
    
    # Calculate totals
    by_category = {}
    grand_total_minutes = 0
    grand_total_cost = 0.0
    
    for r in results:
        cat = r["_id"]
        total_mins = r["total_hours"] * 60 + r["total_minutes"]
        grand_total_minutes += total_mins
        rate = get_rate_for_category(cat)
        cost = round((total_mins / 60) * rate, 2)
        grand_total_cost += cost
        by_category[cat] = {
            "hours": total_mins // 60,
            "minutes": total_mins % 60,
            "total_minutes": total_mins,
            "entry_count": r["entry_count"],
            "hourly_rate": rate,
            "cost": cost
        }
    
    # Get daily breakdown for the period
    daily_pipeline = [
        {"$match": {"date": {"$gte": start_date, "$lt": end_date}}},
        {"$group": {
            "_id": "$date",
            "total_hours": {"$sum": "$hours"},
            "total_minutes": {"$sum": "$minutes"}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    daily_results = list(time_entries.aggregate(daily_pipeline))
    daily_breakdown = []
    for d in daily_results:
        total_mins = d["total_hours"] * 60 + d["total_minutes"]
        daily_breakdown.append({
            "date": d["_id"],
            "hours": total_mins // 60,
            "minutes": total_mins % 60,
            "total_minutes": total_mins
        })
    
    return {
        "period": {
            "start": start_date,
            "end": end_date
        },
        "total": {
            "hours": grand_total_minutes // 60,
            "minutes": grand_total_minutes % 60,
            "total_minutes": grand_total_minutes,
            "total_cost": round(grand_total_cost, 2)
        },
        "by_category": by_category,
        "daily_breakdown": daily_breakdown,
        "entry_count": sum(r["entry_count"] for r in results),
        "rates": CATEGORY_RATES
    }

@router.get("/export")
async def export_entries(
    month: str = Query(..., description="Month in YYYY-MM format")
):
    """Export time entries for a month as Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    year_num, month_num = map(int, month.split("-"))
    start_date = f"{year_num}-{month_num:02d}-01"
    if month_num == 12:
        end_date = f"{year_num + 1}-01-01"
    else:
        end_date = f"{year_num}-{month_num + 1:02d}-01"
    
    # Get entries
    entries = list(time_entries.find({
        "date": {"$gte": start_date, "$lt": end_date}
    }).sort("date", 1))
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Time Tracking {month}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Date", "Hours", "Minutes", "Total (hrs)", "Category", "Rate (£/hr)", "Cost (£)", "Description", "Type"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data rows
    total_minutes = 0
    total_cost = 0.0
    for row, entry in enumerate(entries, 2):
        mins = entry.get("hours", 0) * 60 + entry.get("minutes", 0)
        total_minutes += mins
        category = entry.get("category", "Other")
        rate = get_rate_for_category(category)
        cost = round((mins / 60) * rate, 2)
        total_cost += cost
        
        ws.cell(row=row, column=1, value=entry.get("date", "")).border = border
        ws.cell(row=row, column=2, value=entry.get("hours", 0)).border = border
        ws.cell(row=row, column=3, value=entry.get("minutes", 0)).border = border
        ws.cell(row=row, column=4, value=round(mins / 60, 2)).border = border
        ws.cell(row=row, column=5, value=category).border = border
        ws.cell(row=row, column=6, value=rate).border = border
        ws.cell(row=row, column=7, value=cost).border = border
        ws.cell(row=row, column=8, value=entry.get("description", "")).border = border
        ws.cell(row=row, column=9, value="Auto" if entry.get("auto_tracked") else "Manual").border = border
    
    # Summary row
    summary_row = len(entries) + 3
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=round(total_minutes / 60, 2)).font = Font(bold=True)
    ws.cell(row=summary_row, column=7, value=round(total_cost, 2)).font = Font(bold=True)
    
    # Category summary
    ws.cell(row=summary_row + 2, column=1, value="By Category:").font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=5, value="Rate").font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=6, value="Hours").font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=7, value="Cost").font = Font(bold=True)
    
    # Aggregate by category
    cat_totals = {}
    for entry in entries:
        cat = entry.get("category", "Other")
        mins = entry.get("hours", 0) * 60 + entry.get("minutes", 0)
        cat_totals[cat] = cat_totals.get(cat, 0) + mins
    
    for i, (cat, mins) in enumerate(sorted(cat_totals.items())):
        rate = get_rate_for_category(cat)
        cost = round((mins / 60) * rate, 2)
        ws.cell(row=summary_row + 3 + i, column=1, value=cat)
        ws.cell(row=summary_row + 3 + i, column=5, value=f"£{rate}")
        ws.cell(row=summary_row + 3 + i, column=6, value=round(mins / 60, 2))
        ws.cell(row=summary_row + 3 + i, column=7, value=f"£{cost}")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 10
    
    # Save to bytes
    from fastapi.responses import StreamingResponse
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"time_tracking_{month}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============ Pre-populate Historical Data ============

@router.post("/seed-historical")
async def seed_historical_data():
    """Seed historical time entries based on Emergent development sessions"""
    
    # Check if already seeded
    existing = time_entries.count_documents({})
    if existing > 0:
        return {"message": f"Already have {existing} entries. Delete them first to re-seed.", "seeded": False}
    
    # Historical work estimates - CORRECT TIMELINE: Started ~Feb 2026
    historical_entries = [
        # Week 1 - Mid February 2026 - Initial Setup
        {"date": "2026-02-14", "hours": 4, "minutes": 0, "category": "Development", "description": "Initial project exploration, understanding codebase structure"},
        {"date": "2026-02-15", "hours": 5, "minutes": 30, "category": "Development", "description": "AI Chat personas setup - Hugo, Bob, Margie characters"},
        {"date": "2026-02-16", "hours": 3, "minutes": 0, "category": "App Testing", "description": "Testing AI chat functionality, message flows"},
        {"date": "2026-02-17", "hours": 4, "minutes": 0, "category": "Admin Portal", "description": "Admin portal exploration and staff management"},
        
        # Week 2 - Late February
        {"date": "2026-02-19", "hours": 6, "minutes": 0, "category": "Development", "description": "Safeguarding system review and enhancements"},
        {"date": "2026-02-20", "hours": 5, "minutes": 0, "category": "Development", "description": "AI Safety Classifier implementation with GPT-4o-mini"},
        {"date": "2026-02-21", "hours": 4, "minutes": 30, "category": "Development", "description": "Crisis phrase dataset expansion - 527 phrases"},
        {"date": "2026-02-22", "hours": 3, "minutes": 0, "category": "Documentation", "description": "Safeguarding system documentation"},
        {"date": "2026-02-23", "hours": 4, "minutes": 0, "category": "Staff Portal", "description": "Staff portal case management review"},
        
        # Week 3 - Late February / Early March
        {"date": "2026-02-25", "hours": 5, "minutes": 0, "category": "Development", "description": "Home screen redesign - 2-column grid layout"},
        {"date": "2026-02-26", "hours": 3, "minutes": 30, "category": "Development", "description": "Live chat UI cleanup, header simplification"},
        {"date": "2026-02-27", "hours": 4, "minutes": 0, "category": "LMS Admin", "description": "LMS admin portal review and testing"},
        {"date": "2026-02-28", "hours": 5, "minutes": 0, "category": "Development", "description": "Local conversation storage - encrypted device storage"},
        
        # Week 4 - Early March
        {"date": "2026-03-01", "hours": 4, "minutes": 0, "category": "Support", "description": "Bug fixes - staff status reset, WebSocket issues"},
        {"date": "2026-03-02", "hours": 3, "minutes": 0, "category": "App Testing", "description": "End-to-end testing of chat and support features"},
        {"date": "2026-03-03", "hours": 5, "minutes": 0, "category": "Development", "description": "Events system - Jitsi video integration"},
        {"date": "2026-03-04", "hours": 4, "minutes": 30, "category": "Development", "description": "Jitsi prejoin page fix - disable lobby screen"},
        {"date": "2026-03-05", "hours": 3, "minutes": 0, "category": "Support", "description": "Debugging live support connection issues"},
        
        # Week 5 - Mid March
        {"date": "2026-03-08", "hours": 5, "minutes": 0, "category": "Development", "description": "In-chat voice call implementation"},
        {"date": "2026-03-09", "hours": 4, "minutes": 0, "category": "Development", "description": "WebRTC signaling improvements"},
        {"date": "2026-03-10", "hours": 6, "minutes": 0, "category": "Development", "description": "AI conversation memory - show previous messages"},
        {"date": "2026-03-11", "hours": 3, "minutes": 30, "category": "Support", "description": "Debugging video chat issues, log analysis"},
        {"date": "2026-03-12", "hours": 4, "minutes": 0, "category": "Development", "description": "Jitsi backend config fix - prejoinPageEnabled"},
        
        # This week - Recent
        {"date": "2026-03-13", "hours": 5, "minutes": 0, "category": "Development", "description": "Socket.io event handling for live support"},
        {"date": "2026-03-14", "hours": 4, "minutes": 30, "category": "Support", "description": "Debugging Jitsi video initialization"},
        {"date": "2026-03-15", "hours": 3, "minutes": 0, "category": "Development", "description": "Time tracking system - API and admin portal"},
    ]
    
    now = datetime.now(timezone.utc)
    
    for entry in historical_entries:
        entry["auto_tracked"] = False
        entry["created_at"] = now
        entry["updated_at"] = now
    
    result = time_entries.insert_many(historical_entries)
    
    # Calculate totals
    total_minutes = sum(e["hours"] * 60 + e["minutes"] for e in historical_entries)
    
    return {
        "message": f"Seeded {len(result.inserted_ids)} historical entries",
        "total_hours": round(total_minutes / 60, 1),
        "seeded": True
    }

@router.delete("/clear-all")
async def delete_all_entries(confirm: bool = Query(False)):
    """Delete all time entries (use with caution!)"""
    if not confirm:
        raise HTTPException(status_code=400, detail="Must set confirm=true to delete all entries")
    
    result = time_entries.delete_many({})
    return {"message": f"Deleted {result.deleted_count} entries"}
